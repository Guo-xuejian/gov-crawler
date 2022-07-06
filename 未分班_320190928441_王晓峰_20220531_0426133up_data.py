from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from openpyxl import load_workbook
import re


def get_href(resp):                ###获得超链接，跳转页面列表
    soup = BeautifulSoup(resp, 'html.parser')  # BeautifulSoup中的方法
    num = 1
    list_href=[]
    list_1=[]
    for link in soup.find_all('a'):  # 遍历网页中所有的超链接（a标签）
        if (num % 2 != 0):
            list_href.append(link.get('href'))
        num = num + 1

    list_1 = ['http://njwlwz.gov.cn'+str(i) for i in list_href]   ##http://njwlwz.gov.cn   链接前缀
    list_2 = list_1

    ###删除字符串里的‘..’
    for i in range(0, 20):
        if '..' in list_1[i]:
            list_2[i] = list_1[i].replace('..', '')
    return list_2

def get_content(browser,list_main):
    js = 'window.open("' + str(list_main) + '");'

    # 打开窗口
    browser.execute_script(js)

    # 得到所有的句柄，和主要的句柄
    handle_main = browser.current_window_handle
    handle_all = browser.window_handles
    # 在所有的句柄中找到需要的句柄
    for handle in handle_all:
        if handle != handle_main:
            handle_now = handle
    browser.switch_to.window(handle_now)

    # 在新的窗口中进行爬取
    content_title1 = 'I1'
    content1 = browser.find_element(By.ID, content_title1)
    browser.switch_to.frame(content1)

    # 内容获得
    try:
        ask_name = browser.find_element(By.XPATH,
                                        '//*[@id="ctl00_ContentPlaceHolder1_ctl_MailView1_DetailsView1_Label_UserName"]').text
        ask_time = browser.find_element(By.XPATH,
                                        '//*[@id="ctl00_ContentPlaceHolder1_ctl_MailView1_DetailsView1_Label_AddDate"]').text
        ask_body = browser.find_element(By.XPATH,
                                        '//*[@id="ctl00_ContentPlaceHolder1_ctl_MailView1_DetailsView1_Label_Content"]').text
        answer_name = browser.find_element(By.XPATH,
                                           '//*[@id="ctl00_ContentPlaceHolder1_ctl_MailView1_GridView1_ctl02_Label_UserName"]').text
        answer_time = browser.find_element(By.XPATH,
                                           '//*[@id="ctl00_ContentPlaceHolder1_ctl_MailView1_GridView1_ctl02_Label3"]').text
        answer_body = browser.find_element(By.XPATH,
                                           '//*[@id="ctl00_ContentPlaceHolder1_ctl_MailView1_GridView1"]/tbody/tr[2]/td/table/tbody/tr[2]/td').text
    except:
        ask_name = browser.find_element(By.XPATH,
                                        '//*[@id="ctl00_ContentPlaceHolder1_MailShow1_DetailsView1_Label_UserName"]').text
        ask_time = browser.find_element(By.XPATH,
                                        '//*[@id="ctl00_ContentPlaceHolder1_MailShow1_DetailsView1_Label_AddDate"]').text
        ask_body = browser.find_element(By.XPATH,
                                        '//*[@id="ctl00_ContentPlaceHolder1_MailShow1_DetailsView1_Label_Content"]').text
        answer_name = browser.find_element(By.XPATH,
                                           '//*[@id="ctl00_ContentPlaceHolder1_MailShow1_GridView1_ctl02_Label_UserName"]').text
        answer_time = browser.find_element(By.XPATH,
                                           '//*[@id="ctl00_ContentPlaceHolder1_MailShow1_GridView1_ctl02_Label3"]').text
        answer_body = browser.find_element(By.XPATH,
                                           '//*[@id="ctl00_ContentPlaceHolder1_MailShow1_GridView1_ctl02_Label_AnswerContent"]').text

    dict_1 = {'ask_name': ask_name, 'ask_time': ask_time, 'ask_body': ask_body, 'answer_name': answer_name,
              'answer_time': answer_time, 'answer_body': answer_body}
    # csv_writer_content.writerow(dict_1.values())  ###将内容写入文件里面

    #print(dict_1.values())


    ##关闭handle_now 窗口
    browser.close()
    ## 切换会handle_main 窗口
    browser.switch_to.window(handle_main)

    return dict_1

def main():
    workbook = load_workbook(filename="new_data.xlsx")
    sheet = workbook.active

    hang = sheet.max_row

    for i in range(1, hang + 1):
        sheet.delete_rows(i)
    workbook.save(filename="new_data.xlsx")

    title = (
    'classification', 'time', 'title', 'send_name', 'receive_name', 'answer_name', 'state', 'ask_name', 'ask_time',
    'ask_body', 'answer_name', 'answer_time', 'answer_body')  ##头部标题
    str1 = 'A' + str(1)
    str2 = 'B' + str(1)
    str3 = 'C' + str(1)
    str4 = 'D' + str(1)
    str5 = 'E' + str(1)
    str6 = 'F' + str(1)
    str7 = 'G' + str(1)
    str8 = 'H' + str(1)
    str9 = 'I' + str(1)
    str10 = 'J' + str(1)
    str11 = 'K' + str(1)
    str12 = 'L' + str(1)
    str13 = 'M' + str(1)
    sheet[str1].value = title[0]
    sheet[str2].value = title[1]
    sheet[str3].value = title[2]
    sheet[str4].value = title[3]
    sheet[str5].value = title[4]
    sheet[str6].value = title[5]
    sheet[str7].value = title[6]
    sheet[str8].value = title[7]
    sheet[str9].value = title[8]
    sheet[str10].value = title[9]
    sheet[str11].value = title[10]
    sheet[str12].value = title[11]
    sheet[str13].value = title[12]
    workbook.save(filename="new_data.xlsx")

    workbook_add_1 = load_workbook(filename="header_crawler_data.xlsx")
    sheet_add_1 = workbook_add_1.active

    workbook_add_2 = load_workbook(filename="body_crawler_data.xlsx")
    sheet_add_2 = workbook_add_2.active

    up_data = 0

    l = 1
    browser = webdriver.Chrome()
    url = "http://njwlwz.gov.cn/Mail/Allemail.aspx?MailBox=%u6240%u6709%u4fe1%u7bb1&Manager=%25&ShowTitle=1&MailType=%25&Title=%25&Sender=%25&ID=%25&Mobile=%25&Open=1"

    browser.get(url)

    resp = browser.page_source  ##第一页的页面源代码
    list_main = get_href(resp)      ###将跳转网址放到list_main

    page_number = re.compile(
        r'<span id="ctl00_ContentPlaceHolder1_SmailSearch1_GridView1_ctl23_LabelPageCount">(?P<page_num>.*?)</span>',
        re.S)
    page_number1 = page_number.finditer(resp)  ###页码数量

    for it in page_number1:
        x_all = it.group("page_num")
    print("the total page number is :",x_all)  ###页码数量



    df = pd.read_excel('header_crawler_data.xlsx')  # 这个会直接默认读取到这个Excel的第一个表单

    data = df.loc[0].values  # 0表示第一行 这里读取数据并不包含表头，要注意哦！

    classification_data = data[0]
    time_data =data[1]
    title_data = data[2]
    send_name_data = data[3]
    receive_name_data = data[4]







    for i in range(0, int(x_all)):            ###表示页码数量以及翻页
        element = WebDriverWait(browser, 10).until(
            EC.presence_of_element_located((By.XPATH,
                                            '//*[@id="ctl00_ContentPlaceHolder1_SmailSearch1_GridView1"]/tbody/tr[22]/td/table/tbody/tr/td[9]'))
        )

        obj = re.compile(
            r'<td style="width:35px;">(?P<classification>.*?)</td>.*?<td style="width:80px;">(?P<time>.*?)</td>.*?target="_parent">(?P<title>.*?)</a>.*?<td align="center">(?P<send_name>.*?)</td>.*?'
            r'target="_self">(?P<receive_name>.*?)</a>.*?<td align="center">(?P<answer_name>.*?)</td>.*?<td style="width:55px;">(?P<state>.*?)</td>',
            re.S)
        ##开始匹配
        result = obj.finditer(resp)


        # time.sleep(2)
        number_list = 0
        for it in result:

            dict = it.groupdict()
            #print(dict.values())
            if (dict["classification"] == classification_data and dict["time"] == time_data  and dict["title"] == title_data and dict["send_name"] == send_name_data and dict["receive_name"] == receive_name_data):   ##取出data.csv文件的第一行，最新数据和爬取的最新数据比对



                print("the final up data is ", up_data)

                workbook.save(filename="new_data.xlsx")
                workbook_add_1.save(filename="header_crawler_data.xlsx")
                workbook_add_2.save(filename="body_crawler_data.xlsx")

                print("over")
                browser.quit()
                return


            else:
                up_data += 1

                l += 1

                str1 = 'A' + str(l)
                str2 = 'B' + str(l)
                str3 = 'C' + str(l)
                str4 = 'D' + str(l)
                str5 = 'E' + str(l)
                str6 = 'F' + str(l)
                str7 = 'G' + str(l)
                sheet[str1].value = dict.get('classification')
                sheet[str2].value = dict.get('time')
                sheet[str3].value = dict.get('title')
                sheet[str4].value = dict.get('send_name')
                sheet[str5].value = dict.get('receive_name')
                sheet[str6].value = dict.get('answer_name')
                sheet[str7].value = dict.get('state')


                sheet_add_1.insert_rows(idx=up_data + 1)


                sheet_add_1[str1].value = dict.get('classification')
                sheet_add_1[str2].value = dict.get('time')
                sheet_add_1[str3].value = dict.get('title')
                sheet_add_1[str4].value = dict.get('send_name')
                sheet_add_1[str5].value = dict.get('receive_name')
                sheet_add_1[str6].value = dict.get('answer_name')
                sheet_add_1[str7].value = dict.get('state')



                # dict["time"] = dict["time"].strip()
                # csv_writer.writerow(dict.values())  ###将类型那一页的数据放到文件
                dict_last = get_content(browser, list_main[number_list])
                str8 = 'H' + str(l)
                str9 = 'I' + str(l)
                str10 = 'J' + str(l)
                str11 = 'K' + str(l)
                str12 = 'L' + str(l)
                str13 = 'M' + str(l)



                sheet[str8].value = dict_last.get('ask_name')
                sheet[str9].value = dict_last.get('ask_time')
                sheet[str10].value = dict_last.get('ask_body')
                sheet[str11].value = dict_last.get('answer_name')
                sheet[str12].value = dict_last.get('answer_time')
                sheet[str13].value = dict_last.get('answer_body')



                sheet_add_2.insert_rows(idx=up_data + 1)

                sheet_add_2[str1].value = dict_last.get('ask_name')
                sheet_add_2[str2].value = dict_last.get('ask_time')
                sheet_add_2[str3].value = dict_last.get('ask_body')
                sheet_add_2[str4].value = dict_last.get('answer_name')
                sheet_add_2[str5].value = dict_last.get('answer_time')
                sheet_add_2[str6].value = dict_last.get('answer_body')



                number_list += 1  ##来检查一页中的逐行数据


        try:
            ###点击跳转下一页
            next_xpath = '//*[@id="ctl00_ContentPlaceHolder1_SmailSearch1_GridView1"]/tbody/tr[22]/td/table/tbody/tr/td[4]'
            next_page = browser.find_element(By.XPATH, next_xpath)
            ###print(next_page.text)
            # time.sleep(2)
            next_page.click()

            resp = browser.page_source
        except:
            workbook = load_workbook(filename="new_data.xlsx")

main()

